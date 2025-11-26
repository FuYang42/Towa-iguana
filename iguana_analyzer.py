#!/usr/bin/env python3
"""
Iguana芯片寄存器数据分析工具
从PCAP文件中提取并分析12个Iguana芯片的寄存器数据
"""

from scapy.all import rdpcap
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os


class IguanaRegisterAnalyzer:
    """Iguana芯片寄存器数据分析器"""

    NUM_CHIPS = 12
    REGISTERS_PER_CHIP = 4
    BYTES_PER_REGISTER = 4
    BYTES_PER_CHIP = REGISTERS_PER_CHIP * BYTES_PER_REGISTER  # 16字节
    TOTAL_DATA_SIZE = NUM_CHIPS * BYTES_PER_CHIP  # 192字节
    TARGET_PACKET_LENGTH = 738

    def __init__(self, pcap_file1, pcap_file2=None):
        """
        初始化分析器

        Args:
            pcap_file1: 第一个PCAP文件路径
            pcap_file2: 第二个PCAP文件路径（可选，用于对比）
        """
        self.pcap_file1 = pcap_file1
        self.pcap_file2 = pcap_file2
        self.pcap_file = pcap_file1  # 保持向后兼容

        # 第一个文件的数据
        self.packets_data = []  # 存储所有提取的192字节数据
        self.register_values = defaultdict(list)  # {(chip_id, reg_id): [values...]}

        # 第二个文件的数据
        self.packets_data2 = []
        self.register_values2 = defaultdict(list)

    def extract_packets(self, file_path=None, packets_data_list=None):
        """从PCAP文件中提取长度为738字节的数据包"""
        if file_path is None:
            file_path = self.pcap_file
        if packets_data_list is None:
            packets_data_list = self.packets_data

        print(f"Reading PCAP file: {file_path}")
        try:
            packets = rdpcap(file_path)
        except Exception as e:
            print(f"Error: Unable to read PCAP file - {e}")
            sys.exit(1)

        # 筛选长度为738的数据包
        filtered_count = 0
        for pkt in packets:
            if len(pkt) == self.TARGET_PACKET_LENGTH:
                # 提取最后192字节
                raw_data = bytes(pkt)
                iguana_data = raw_data[-self.TOTAL_DATA_SIZE:]
                packets_data_list.append(iguana_data)
                filtered_count += 1

        print(f"Found {filtered_count} packets of 738 bytes, extracted {filtered_count} register data sets\n")
        return filtered_count

    def reverse_bytes(self, data):
        """
        反转字节序（小端转大端或大端转小端）

        Args:
            data: 4字节的bytes对象

        Returns:
            bytes: 反转后的字节序
        """
        return data[::-1]

    def bytes_to_binary(self, data):
        """
        将字节数据转换为二进制字符串

        Args:
            data: bytes对象

        Returns:
            str: 二进制字符串（例如：'00110101...'）
        """
        return ''.join(format(byte, '08b') for byte in data)

    def parse_register_data(self, packets_data_list=None, register_values_dict=None):
        """解析所有数据包中的寄存器数据"""
        if packets_data_list is None:
            packets_data_list = self.packets_data
        if register_values_dict is None:
            register_values_dict = self.register_values

        for pkt_idx, data in enumerate(packets_data_list):
            for chip_id in range(self.NUM_CHIPS):
                chip_offset = chip_id * self.BYTES_PER_CHIP

                for reg_id in range(self.REGISTERS_PER_CHIP):
                    reg_offset = chip_offset + reg_id * self.BYTES_PER_REGISTER

                    # 提取4字节寄存器数据
                    reg_data_raw = data[reg_offset:reg_offset + self.BYTES_PER_REGISTER]

                    # 反转字节序
                    reg_data_reversed = self.reverse_bytes(reg_data_raw)

                    # 转换为二进制
                    binary_value = self.bytes_to_binary(reg_data_reversed)

                    # 存储该寄存器的值
                    key = (chip_id, reg_id)
                    register_values_dict[key].append(binary_value)

    def compare_and_report(self):
        """比较所有数据包中相同芯片相同寄存器的值，并生成报告"""
        print("=" * 80)
        print("Register Value Analysis Report")
        print("=" * 80)

        all_identical = True
        inconsistent_regs = []

        # 存储每个寄存器在所有芯片中的值（用于跨芯片对比）
        register_across_chips = {}

        for chip_id in range(self.NUM_CHIPS):
            for reg_id in range(self.REGISTERS_PER_CHIP):
                key = (chip_id, reg_id)
                values = self.register_values[key]

                # 获取唯一值
                unique_values = list(set(values))

                if len(unique_values) == 1:
                    # 所有值相同
                    value = unique_values[0]
                    print(f"CHIP{chip_id:2d} Reg{reg_id}  ✓  "
                          f"0x{int(value, 2):08X}  "
                          f"{value}")

                    # 存储该寄存器的值用于跨芯片对比
                    if reg_id not in register_across_chips:
                        register_across_chips[reg_id] = []
                    register_across_chips[reg_id].append((chip_id, value))
                else:
                    # 发现不一致
                    all_identical = False
                    inconsistent_regs.append((chip_id, reg_id))
                    print(f"CHIP{chip_id:2d} Reg{reg_id}  ✗  Inconsistent!")

                    # 统计每个值出现的次数
                    value_counts = {}
                    for val in values:
                        value_counts[val] = value_counts.get(val, 0) + 1

                    # 按出现次数排序
                    sorted_values = sorted(value_counts.items(),
                                         key=lambda x: x[1],
                                         reverse=True)

                    for idx, (val, count) in enumerate(sorted_values, 1):
                        packet_indices = [i+1 for i, v in enumerate(values) if v == val]
                        print(f"            Value{idx} ({count}x): "
                              f"0x{int(val, 2):08X}  {val}  "
                              f"Packets{packet_indices}")

        print("=" * 80)
        if all_identical:
            print("✓ All register values are consistent across packets")
        else:
            print(f"✗ Found {len(inconsistent_regs)} inconsistent registers across packets")
        print("=" * 80)

        # 跨芯片对比
        self.compare_across_chips(register_across_chips)

        return all_identical

    def compare_across_chips(self, register_across_chips):
        """对比不同芯片的相同寄存器值"""
        print("\n" + "=" * 80)
        print("Cross-Chip Register Comparison")
        print("=" * 80)

        cross_chip_inconsistent = False

        for reg_id in range(self.REGISTERS_PER_CHIP):
            if reg_id not in register_across_chips:
                continue

            chip_values = register_across_chips[reg_id]

            # 获取所有不同的值
            unique_values = {}
            for chip_id, value in chip_values:
                if value not in unique_values:
                    unique_values[value] = []
                unique_values[value].append(chip_id)

            if len(unique_values) == 1:
                # 所有芯片的该寄存器值相同
                value = list(unique_values.keys())[0]
                print(f"\nReg{reg_id}  ✓  All chips consistent")
                print(f"  Value: 0x{int(value, 2):08X}  {value}")
            else:
                # 发现不同芯片的该寄存器值不一致
                cross_chip_inconsistent = True
                print(f"\nReg{reg_id}  ✗  Inconsistent across chips!")

                # 按芯片数量排序
                sorted_values = sorted(unique_values.items(),
                                     key=lambda x: len(x[1]),
                                     reverse=True)

                for idx, (value, chips) in enumerate(sorted_values, 1):
                    print(f"  Value{idx} (CHIP {len(chips)}x): 0x{int(value, 2):08X}  {value}")
                    print(f"    Chips: {chips}")

        print("=" * 80)
        if cross_chip_inconsistent:
            print("✗ Found inconsistencies across chips")
        else:
            print("✓ All registers are consistent across all chips")
        print("=" * 80)

    def compare_two_files(self):
        """对比两个PCAP文件的寄存器数据并生成对比报告"""
        print("\n" + "=" * 80)
        print("REGISTER VALUE COMPARISON BETWEEN TWO FILES")
        print("=" * 80)

        # 统计信息
        total_differences = 0
        identical_count = 0

        # 对比每个芯片的每个寄存器
        for chip_id in range(self.NUM_CHIPS):
            chip_has_diff = False
            chip_diff_details = []

            for reg_id in range(self.REGISTERS_PER_CHIP):
                key = (chip_id, reg_id)

                # 获取两个文件中该寄存器的值
                values1 = self.register_values.get(key, [])
                values2 = self.register_values2.get(key, [])

                if not values1 or not values2:
                    continue

                # 获取第一个值（代表性值）
                value1 = values1[0] if values1 else None
                value2 = values2[0] if values2 else None

                if value1 != value2:
                    chip_has_diff = True
                    total_differences += 1

                    # 计算差异位
                    diff_bits = []
                    for i, (bit1, bit2) in enumerate(zip(value1, value2)):
                        if bit1 != bit2:
                            diff_bits.append(i)

                    chip_diff_details.append({
                        'reg_id': reg_id,
                        'value1': value1,
                        'value2': value2,
                        'diff_bits': diff_bits
                    })
                else:
                    identical_count += 1

            # 输出该芯片的对比结果
            if chip_has_diff:
                print(f"\nCHIP{chip_id:2d}  ✗  Differences found:")
                for detail in chip_diff_details:
                    reg_id = detail['reg_id']
                    value1 = detail['value1']
                    value2 = detail['value2']
                    diff_bits = detail['diff_bits']

                    print(f"  Reg{reg_id}:")
                    print(f"    File1: 0x{int(value1, 2):08X}  {value1}")
                    print(f"    File2: 0x{int(value2, 2):08X}  {value2}")
                    if diff_bits:
                        print(f"    Diff at bits: {diff_bits} (total: {len(diff_bits)} bits)")
            else:
                print(f"\nCHIP{chip_id:2d}  ✓  All registers identical")

        # 汇总统计
        print("\n" + "=" * 80)
        print("COMPARISON SUMMARY")
        print("=" * 80)
        total_registers = self.NUM_CHIPS * self.REGISTERS_PER_CHIP
        print(f"Total registers compared: {total_registers}")
        print(f"Identical registers: {identical_count}")
        print(f"Different registers: {total_differences}")

        if total_differences == 0:
            print("\n✓ All register values are IDENTICAL in both files!")
        else:
            percentage = (total_differences / total_registers) * 100
            print(f"\n✗ Found differences in {percentage:.1f}% of registers")

        print("=" * 80)

    def export_to_excel_single(self, output_file):
        """导出单文件分析结果到Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Register Analysis"

        # 设置表头
        headers = ["ChipID", "RegID", "Value (Hex)", "Value (Binary)"]
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        for chip_id in range(self.NUM_CHIPS):
            for reg_id in range(self.REGISTERS_PER_CHIP):
                key = (chip_id, reg_id)
                values = self.register_values.get(key, [])

                if values:
                    # 使用第一个值作为代表值
                    binary_value = values[0]
                    hex_value = f"0x{int(binary_value, 2):08X}"

                    ws.append([chip_id, reg_id, hex_value, binary_value])

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40

        # 保存文件
        wb.save(output_file)
        print(f"\n✓ Excel file saved: {output_file}")

    def run_analysis(self):
        """运行完整的分析流程"""
        # 提取数据包
        count = self.extract_packets()

        if count == 0:
            print(f"No packets of {self.TARGET_PACKET_LENGTH} bytes found")
            return

        # 解析寄存器数据
        print("Parsing register data...\n")
        self.parse_register_data()

        # 对比并生成报告
        self.compare_and_report()

        # 导出到Excel
        base_name = os.path.splitext(os.path.basename(self.pcap_file))[0]
        excel_file = f"{base_name}_analysis.xlsx"
        self.export_to_excel_single(excel_file)

    def export_to_excel_comparison(self, output_file):
        """导出双文件对比结果到Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Register Comparison"

        # 获取文件名（不含扩展名）
        file1_name = os.path.splitext(os.path.basename(self.pcap_file1))[0]
        file2_name = os.path.splitext(os.path.basename(self.pcap_file2))[0]

        # 设置表头，使用实际的文件名
        headers = ["ChipID", "RegID",
                   f"{file1_name} (Hex)", f"{file1_name} (Binary)",
                   f"{file2_name} (Hex)", f"{file2_name} (Binary)",
                   "Status"]
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        for chip_id in range(self.NUM_CHIPS):
            for reg_id in range(self.REGISTERS_PER_CHIP):
                key = (chip_id, reg_id)

                # 获取两个文件中该寄存器的值
                values1 = self.register_values.get(key, [])
                values2 = self.register_values2.get(key, [])

                if values1 and values2:
                    binary_value1 = values1[0]
                    binary_value2 = values2[0]
                    hex_value1 = f"0x{int(binary_value1, 2):08X}"
                    hex_value2 = f"0x{int(binary_value2, 2):08X}"

                    # 判断是否相同
                    status = "Identical" if binary_value1 == binary_value2 else "Different"

                    row = [chip_id, reg_id, hex_value1, binary_value1, hex_value2, binary_value2, status]
                    ws.append(row)

                    # 如果不同，高亮显示该行
                    if status == "Different":
                        row_num = ws.max_row
                        diff_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                        for cell in ws[row_num]:
                            cell.fill = diff_fill

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 12

        # 保存文件
        wb.save(output_file)
        print(f"\n✓ Excel comparison file saved: {output_file}")

    def run_comparison(self):
        """运行双文件对比分析流程"""
        print("\n" + "=" * 80)
        print("DUAL FILE COMPARISON MODE")
        print("=" * 80)

        # 提取第一个文件的数据包
        print(f"\n[FILE 1: {self.pcap_file1}]")
        count1 = self.extract_packets(self.pcap_file1, self.packets_data)
        if count1 == 0:
            print(f"No packets of {self.TARGET_PACKET_LENGTH} bytes found in file 1")
            return

        # 提取第二个文件的数据包
        print(f"\n[FILE 2: {self.pcap_file2}]")
        count2 = self.extract_packets(self.pcap_file2, self.packets_data2)
        if count2 == 0:
            print(f"No packets of {self.TARGET_PACKET_LENGTH} bytes found in file 2")
            return

        # 解析两个文件的寄存器数据
        print("Parsing register data from both files...\n")
        self.parse_register_data(self.packets_data, self.register_values)
        self.parse_register_data(self.packets_data2, self.register_values2)

        # 对比两个文件的寄存器数据
        self.compare_two_files()

        # 导出到Excel
        base_name1 = os.path.splitext(os.path.basename(self.pcap_file1))[0]
        base_name2 = os.path.splitext(os.path.basename(self.pcap_file2))[0]
        excel_file = f"{base_name1}_vs_{base_name2}_comparison.xlsx"
        self.export_to_excel_comparison(excel_file)


def main():
    """主函数"""
    print("=" * 80)
    print("Iguana Chip Register Data Analysis Tool")
    print("=" * 80)

    # 检查命令行参数
    if len(sys.argv) < 2:
        print("\nUsage: python iguana_analyzer.py <PCAP_file_path> [PCAP_file_path_2]")
        print("\nExamples:")
        print("  Single file analysis:")
        print("    python iguana_analyzer.py test_tool.pcap")
        print("    python iguana_analyzer.py C:\\Users\\FuYou\\Desktop\\data.pcap")
        print("\n  Dual file comparison:")
        print("    python iguana_analyzer.py file1.pcap file2.pcap")
        print("    python iguana_analyzer.py before.pcap after.pcap")
        sys.exit(1)

    pcap_file1 = sys.argv[1]
    pcap_file2 = sys.argv[2] if len(sys.argv) >= 3 else None

    if pcap_file2:
        # 双文件对比模式
        analyzer = IguanaRegisterAnalyzer(pcap_file1, pcap_file2)
        analyzer.run_comparison()
    else:
        # 单文件分析模式
        analyzer = IguanaRegisterAnalyzer(pcap_file1)
        analyzer.run_analysis()


if __name__ == "__main__":
    main()
